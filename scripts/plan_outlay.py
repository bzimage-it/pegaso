#!/usr/bin/env python3
"""
Script per generare estratto conto di pianificazione finanziaria
a partire da configurazione JSON con accumuli e spese.
"""

import json
import csv
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import argparse
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule

def parse_amount(amount_str):
    """Converte una stringa con formato italiano (es: '139,60') in float."""
    return float(amount_str.replace(',', '.'))

def format_amount(amount):
    """Formatta un float in stringa con formato italiano."""
    if amount == 0:
        return "0"
    return f"{amount:.2f}".replace('.', ',')

def parse_date(date_str):
    """Converte una stringa data in oggetto datetime."""
    return datetime.strptime(date_str, '%Y-%m-%d')


def generate_plan_events(plan, data_inizio, durata_anni):
    """Genera tutti gli eventi del piano (accumuli e spese) nel periodo specificato.
    Gestisce l'aggiornamento degli eventi ricorrenti: quando un evento con la stessa 
    descrizione riappare, quello nuovo sostituisce il precedente per le ricorrenze future.
    Per colonne tipo Accumulo in estratto: is_original_date è True sulla prima occorrenza
    di ogni segmento che cade su o dopo data_inizio (stessa griglia modulare), non sulla
    prima occorrenza assoluta se questa è precedente a data_inizio.
    """
    events = []
    data_inizio_dt = parse_date(data_inizio)
    data_fine = data_inizio_dt + relativedelta(years=durata_anni)
    
    # Ordina gli eventi del piano per data per processarli cronologicamente
    plan_sorted = sorted(plan, key=lambda x: parse_date(x['data']))
    
    # Dizionario per tracciare l'ultima configurazione di ogni descrizione
    eventi_attivi = {}
    
    # Prima passata: identifica le configurazioni attive per ogni periodo
    for evento in plan_sorted:
        # Ignora l'evento se ha skip=true
        if evento.get('skip', False):
            continue
            
        descrizione = evento['descrizione'].upper()  # Trasforma in uppercase
        data_evento = parse_date(evento['data'])
        importo = parse_amount(evento['importo'])
        
        if descrizione not in eventi_attivi:
            eventi_attivi[descrizione] = []
        
        eventi_attivi[descrizione].append({
            'data_inizio': data_evento,
            'importo': importo,
            'tipo': evento['tipo'],
            'frequenza': evento.get('frequenza', 1),  # Default 1 se non specificato
            'descrizione': descrizione
        })
    
    # Seconda passata: genera gli eventi considerando i periodi di validità
    for descrizione, configurazioni in eventi_attivi.items():
        for i, config in enumerate(configurazioni):
            data_config_inizio = config['data_inizio']
            
            # Determina quando questa configurazione termina (quando inizia la prossima)
            if i < len(configurazioni) - 1:
                data_config_fine = configurazioni[i + 1]['data_inizio']
            else:
                data_config_fine = data_fine
            
            # Genera eventi per questa configurazione
            frequenza = config['frequenza']
            
            if config['tipo'] == 'una tantum':
                # Applica solo una volta
                if data_config_inizio < data_config_fine:
                    # Determina il tipo in base al segno dell'importo
                    tipo_evento = 'accumulo' if config['importo'] > 0 else 'spesa'
                    events.append({
                        'data': data_config_inizio.strftime('%Y-%m-%d'),
                        'descrizione': descrizione,
                        'tipo': tipo_evento,
                        'importo': config['importo'],
                        'is_original_date': data_config_inizio >= data_inizio_dt
                    })
            
            elif config['tipo'] == 'settimanale':
                # Applica ogni N settimane fino alla prossima configurazione o alla fine
                current_date = data_config_inizio
                tipo_evento = 'accumulo' if config['importo'] > 0 else 'spesa'
                seen_in_estratto_window = False
                while current_date < data_config_fine:
                    in_window = current_date >= data_inizio_dt
                    is_original_date = in_window and not seen_in_estratto_window
                    if in_window:
                        seen_in_estratto_window = True
                    events.append({
                        'data': current_date.strftime('%Y-%m-%d'),
                        'descrizione': descrizione,
                        'tipo': tipo_evento,
                        'importo': config['importo'],
                        'is_original_date': is_original_date
                    })
                    current_date += timedelta(weeks=frequenza)
            
            elif config['tipo'] == 'mensile':
                # Applica ogni N mesi fino alla prossima configurazione o alla fine
                current_date = data_config_inizio
                tipo_evento = 'accumulo' if config['importo'] > 0 else 'spesa'
                seen_in_estratto_window = False
                while current_date < data_config_fine:
                    in_window = current_date >= data_inizio_dt
                    is_original_date = in_window and not seen_in_estratto_window
                    if in_window:
                        seen_in_estratto_window = True
                    events.append({
                        'data': current_date.strftime('%Y-%m-%d'),
                        'descrizione': descrizione,
                        'tipo': tipo_evento,
                        'importo': config['importo'],
                        'is_original_date': is_original_date
                    })
                    current_date += relativedelta(months=frequenza)
            
            elif config['tipo'] == 'annuale':
                # Applica ogni N anni fino alla prossima configurazione o alla fine
                current_date = data_config_inizio
                tipo_evento = 'accumulo' if config['importo'] > 0 else 'spesa'
                seen_in_estratto_window = False
                while current_date < data_config_fine:
                    in_window = current_date >= data_inizio_dt
                    is_original_date = in_window and not seen_in_estratto_window
                    if in_window:
                        seen_in_estratto_window = True
                    events.append({
                        'data': current_date.strftime('%Y-%m-%d'),
                        'descrizione': descrizione,
                        'tipo': tipo_evento,
                        'importo': config['importo'],
                        'is_original_date': is_original_date
                    })
                    current_date += relativedelta(years=frequenza)
    
    return events

def generate_estratto_conto(config_file, output_file):
    """Genera l'estratto conto completo."""
    
    # Carica configurazione
    with open(config_file, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    saldo_iniziale = parse_amount(config['saldo_iniziale'])
    data_inizio = config['data_inizio']
    durata_anni = config['durata_anni']
    plan = config.get('plan', [])
    
    # Genera eventi
    all_events = generate_plan_events(plan, data_inizio, durata_anni)
    
    # Filtra eventi che sono dopo o uguali alla data di inizio
    data_inizio_dt = parse_date(data_inizio)
    all_events = [event for event in all_events if parse_date(event['data']) >= data_inizio_dt]
    
    # Ordina per data
    all_events.sort(key=lambda x: parse_date(x['data']))
    
    # Crea l'estratto conto
    estratto = []
    
    # Aggiungi evento iniziale
    estratto.append({
        'data': data_inizio,
        'cosa': 'SALDO ATTUALE',
        'Accumulo': format_amount(saldo_iniziale),  # Saldo iniziale nella colonna Accumulo
        'Importo': 'FORMULA_IMPORTO_INIZIALE',      # Sarà sostituito con =C2
        'Saldo': 'FORMULA_SALDO_INIZIALE'          # Sarà sostituito con =D2
    })
    
    saldo_corrente = saldo_iniziale
    
    # Processa tutti gli eventi
    for event in all_events:
        saldo_corrente += event['importo']
        
        # Gestisci accumuli e spese diversamente
        if event['importo'] > 0:  # Accumulo
            # Valorizza la colonna Accumulo solo se è la data originale
            accumulo_val = format_amount(event['importo']) if event.get('is_original_date', False) else ''
            estratto.append({
                'data': event['data'],
                'cosa': event['descrizione'],
                'Accumulo': accumulo_val,
                'Importo': 'FORMULA_ACCUMULO',  # Sarà sostituito con formula Excel
                'Saldo': format_amount(saldo_corrente)
            })
        else:  # Spesa
            estratto.append({
                'data': event['data'],
                'cosa': event['descrizione'],
                'Accumulo': '',
                'Importo': format_amount(event['importo']),
                'Saldo': format_amount(saldo_corrente)
            })
    
    # Scrivi Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estratto Conto"
    
    # Scrivi header
    headers = ['Data', 'Cosa', 'Accumulo', 'Importo', 'Saldo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # Formattazione header: grassetto, size 14
        cell.font = Font(bold=True, size=14)
        
        # Allineamento: A,B a sinistra, C,D,E a destra
        if col <= 2:  # Colonne A, B (Data, Cosa)
            cell.alignment = Alignment(horizontal='left')
        else:  # Colonne C, D, E (Accumulo, Importo, Saldo)
            cell.alignment = Alignment(horizontal='right')
    
    # Scrivi dati
    for row_idx, row_data in enumerate(estratto, 2):
        ws.cell(row=row_idx, column=1, value=row_data['data'])
        ws.cell(row=row_idx, column=2, value=row_data['cosa'])
        
        # Accumulo - converti da stringa a numero se non vuoto
        accumulo_val = row_data['Accumulo']
        if accumulo_val and accumulo_val != '':
            ws.cell(row=row_idx, column=3, value=parse_amount(accumulo_val))
        else:
            ws.cell(row=row_idx, column=3, value='')
        
        # Importo - gestisci formula per accumuli o valore per spese
        importo_val = row_data['Importo']
        if importo_val == 'FORMULA_IMPORTO_INIZIALE':
            # Prima riga: formula =C2 per prendere il valore dalla colonna Accumulo
            ws.cell(row=row_idx, column=4, value="=C2")
        elif importo_val == 'FORMULA_ACCUMULO':
            # Accumulo: inserisci sempre la formula per cercare ultimo valore sopra
            if row_idx == 2:  # Prima riga, usa il valore in Accumulo se presente
                if accumulo_val and accumulo_val != '':
                    ws.cell(row=row_idx, column=4, value=parse_amount(accumulo_val))
                else:
                    ws.cell(row=row_idx, column=4, value='')
            else:
                formula = f"=IF(C{row_idx}<>\"\",C{row_idx},INDEX(C$2:C{row_idx},LOOKUP(2,1/(C$2:C{row_idx}<>\"\"),ROW(C$2:C{row_idx})-1)))"
                ws.cell(row=row_idx, column=4, value=formula)
        elif importo_val and importo_val != '':
            # Spesa: inserisci il valore direttamente
            ws.cell(row=row_idx, column=4, value=parse_amount(importo_val))
        else:
            ws.cell(row=row_idx, column=4, value='')
        
        # Saldo - formula per calcolo automatico
        saldo_val = row_data['Saldo']
        if saldo_val == 'FORMULA_SALDO_INIZIALE':
            # Prima riga: formula =D2 per prendere il valore dalla colonna Importo
            ws.cell(row=row_idx, column=5, value="=D2")
        elif row_idx == 2:  # Caso di backup se non è formula speciale
            ws.cell(row=row_idx, column=5, value=parse_amount(saldo_val))
        else:
            # Formula: saldo precedente + importo (dalla colonna D)
            formula = f"=E{row_idx-1}+IF(D{row_idx}=\"\",0,D{row_idx})"
            ws.cell(row=row_idx, column=5, value=formula)
    
    # Imposta larghezza colonne
    # Colonna A (Data): larghezza per date YYYY-MM-DD
    ws.column_dimensions['A'].width = 12
    
    # Colonna B (Cosa): larghezza per descrizioni
    ws.column_dimensions['B'].width = 25
    
    # Colonna C (Accumulo): larghezza per numeri con formato italiano
    ws.column_dimensions['C'].width = 15
    
    # Colonna D (Importo): larghezza per numeri con formato italiano
    ws.column_dimensions['D'].width = 15
    
    # Colonna E (Saldo): larghezza per numeri con formato italiano
    ws.column_dimensions['E'].width = 15
    
    # Applica formato numerico con 2 decimali alle colonne numeriche
    # Formato: 0.00 per mostrare sempre 2 decimali
    numero_format = '0.00'
    
    # Applica a tutte le celle delle colonne C, D, E (Accumulo, Importo, Saldo)
    for row in range(2, len(estratto) + 2):  # Dalla riga 2 in poi (esclude header)
        # Colonna C (Accumulo)
        ws.cell(row=row, column=3).number_format = numero_format
        # Colonna D (Importo) 
        ws.cell(row=row, column=4).number_format = numero_format
        # Colonna E (Saldo)
        ws.cell(row=row, column=5).number_format = numero_format
    
    # Applica formattazione arancione grassetto per tutta la colonna Accumulo (C)
    orange_font = Font(color="FF8C00", bold=True)  # Arancione scuro
    for row in range(1, len(estratto) + 2):  # Include anche l'header
        ws.cell(row=row, column=3).font = orange_font
    
    # Applica formattazione condizionale per saldi negativi
    # Rosso grassetto per valori < 0 nella colonna E (Saldo)
    red_font = Font(color="FF0000", bold=True)
    rule = CellIsRule(operator='lessThan', formula=['0'], font=red_font)
    ws.conditional_formatting.add(f'E2:E{len(estratto) + 1}', rule)
    
    # Salva il file
    wb.save(output_file)
    
    print(f"Estratto conto generato in: {output_file}")
    print(f"Numero di eventi: {len(estratto)}")
    print(f"Saldo finale: {format_amount(saldo_corrente)}")

def main():
    parser = argparse.ArgumentParser(description='Genera estratto conto da configurazione JSON')
    parser.add_argument('config_file', help='File JSON di configurazione')
    parser.add_argument('-o', '--output', default=None,
                       help='File Excel di output (default: stesso percorso del config con estensione .xlsx)')
    
    args = parser.parse_args()
    output_file = args.output or str(Path(args.config_file).with_suffix('.xlsx'))
    
    try:
        generate_estratto_conto(args.config_file, output_file)
    except Exception as e:
        print(f"Errore: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
